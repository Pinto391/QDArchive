"""
classification/generate_xlsx_report.py
----------------------------------------
Generates a submission-ready XLSX workbook of the Part 2 classification
results for the "Seeding QDArchive" project (ISIC Rev. 5 classification of
every project and every downloaded file).

Usage:
    python classification/generate_xlsx_report.py working.db --out 23123639_SQ26_Part2_Classification.xlsx

Design notes:
  - Two sheets ("Project Classifications", "File Classifications") hold the
    raw classification data one row per project / per file.
  - The "Overview", "Section Distribution", "Division Distribution", and
    "Tags" sheets are all computed with live formulas (COUNTA / COUNTIFS /
    AVERAGEIFS) against those raw sheets, so re-opening the file after
    editing the raw rows recalculates the summary — nothing is a frozen
    Python-computed literal.
"""

import argparse
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

STUDENT_NAME = "Almas Ali Pinto"
STUDENT_ID = "23123639"

FONT_NAME = "Arial"
ACCENT = "1F4E5F"
ACCENT_LIGHT = "EAF1F3"
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=ACCENT)
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, color="555555")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Fixed, generous upper bound for formula ranges (COUNTA/COUNTIFS/AVERAGEIFS)
# so summary formulas keep working if more rows are appended later without
# needing to be rewritten.
MAX_ROW = 100000


def _fetch(db_path: str):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    projects = conn.execute(
        """
        SELECT p.id AS project_id, p.title, r.name AS repository, p.download_method,
               p.doi, p.project_url,
               c.isic_section, c.section_name, c.isic_division, c.division_name,
               c.confidence, c.method
        FROM PROJECTS p
        JOIN REPOSITORIES r ON p.repository_id = r.id
        LEFT JOIN CLASSIFICATIONS c ON c.project_id = p.id
        ORDER BY p.id
        """
    ).fetchall()

    files = conn.execute(
        """
        SELECT f.id AS file_id, f.project_id, f.file_name, f.file_type,
               fc.isic_section, fc.isic_division, fc.division_name,
               fc.confidence, fc.method
        FROM FILES f
        JOIN FILE_CLASSIFICATIONS fc ON fc.file_id = f.id
        ORDER BY f.project_id, f.id
        """
    ).fetchall()

    project_tags = conn.execute(
        "SELECT project_id, tag FROM TAGS ORDER BY tag, project_id"
    ).fetchall()

    isic_lookup = conn.execute(
        """
        SELECT DISTINCT isic_section, section_name, isic_division, division_name
        FROM CLASSIFICATIONS
        ORDER BY isic_section, isic_division
        """
    ).fetchall()

    conn.close()
    return (
        [dict(r) for r in projects],
        [dict(r) for r in files],
        [dict(r) for r in project_tags],
        [dict(r) for r in isic_lookup],
    )


def _style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _write_table(ws, headers, rows, start_row=1, table_name=None, col_widths=None):
    n_cols = len(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    _style_header_row(ws, start_row, n_cols)

    for r, row_data in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER

    if col_widths:
        for c, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    if table_name and rows:
        last_row = start_row + len(rows)
        last_col = get_column_letter(n_cols)
        ref = f"{get_column_letter(1)}{start_row}:{last_col}{last_row}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
        ws.add_table(tbl)


def build_workbook(db_path: str, out_path: str):
    projects, files, project_tags, isic_lookup = _fetch(db_path)

    wb = Workbook()

    # ── Sheet: Project Classifications ─────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Project Classifications"
    proj_headers = [
        "project_id", "title", "repository", "download_method", "doi",
        "isic_section", "section_name", "isic_division", "division_name",
        "confidence", "method",
    ]
    proj_rows = [
        [
            p["project_id"], p["title"], p["repository"], p["download_method"], p["doi"],
            p["isic_section"], p["section_name"], p["isic_division"], p["division_name"],
            p["confidence"], p["method"],
        ]
        for p in projects
    ]
    _write_table(
        ws_proj, proj_headers, proj_rows, table_name="ProjectClassifications",
        col_widths=[10, 45, 10, 15, 20, 10, 42, 12, 42, 11, 20],
    )
    for r in range(2, len(proj_rows) + 2):
        ws_proj.cell(row=r, column=10).number_format = "0.00"

    # ── Sheet: File Classifications ────────────────────────────────────
    ws_file = wb.create_sheet("File Classifications")
    file_headers = [
        "file_id", "project_id", "project_title", "file_name", "file_type",
        "isic_section", "isic_division", "division_name", "confidence", "method",
    ]
    file_rows = []
    for f in files:
        file_rows.append([
            f["file_id"], f["project_id"],
            # project_title filled via formula below (INDEX/MATCH), placeholder here
            None,
            f["file_name"], f["file_type"],
            f["isic_section"], f["isic_division"], f["division_name"],
            f["confidence"], f["method"],
        ])
    _write_table(
        ws_file, file_headers, file_rows, table_name="FileClassifications",
        col_widths=[10, 11, 45, 40, 11, 11, 12, 42, 11, 20],
    )
    for r in range(2, len(file_rows) + 2):
        ws_file.cell(row=r, column=9).number_format = "0.00"
    # project_title: live lookup against Project Classifications sheet (INDEX/MATCH,
    # not XLOOKUP, for compatibility)
    for r in range(2, len(file_rows) + 2):
        ws_file.cell(row=r, column=3).value = (
            f"=IFERROR(INDEX('Project Classifications'!$B$2:$B${len(proj_rows)+1},"
            f"MATCH(B{r},'Project Classifications'!$A$2:$A${len(proj_rows)+1},0)),\"\")"
        )
        ws_file.cell(row=r, column=3).font = BODY_FONT
        ws_file.cell(row=r, column=3).border = BORDER

    # ── Sheet: Project-Tags (raw mapping) ──────────────────────────────
    ws_tags_raw = wb.create_sheet("Project-Tags")
    _write_table(
        ws_tags_raw, ["project_id", "tag"],
        [[t["project_id"], t["tag"]] for t in project_tags],
        table_name="ProjectTags", col_widths=[11, 35],
    )

    # ── Sheet: Tags (aggregated, live COUNTIF against Project-Tags) ───
    ws_tags = wb.create_sheet("Tags")
    distinct_tags = sorted({t["tag"] for t in project_tags})
    tag_headers = ["tag", "project_count"]
    _style_header_row(ws_tags, 1, 2)
    ws_tags.cell(row=1, column=1, value="tag")
    ws_tags.cell(row=1, column=2, value="project_count")
    for i, tag in enumerate(distinct_tags, start=2):
        ws_tags.cell(row=i, column=1, value=tag).font = BODY_FONT
        ws_tags.cell(row=i, column=1).border = BORDER
        cnt_cell = ws_tags.cell(
            row=i, column=2,
            value=f"=COUNTIF('Project-Tags'!$B$2:$B${MAX_ROW},A{i})",
        )
        cnt_cell.font = BODY_FONT
        cnt_cell.border = BORDER
    ws_tags.column_dimensions["A"].width = 35
    ws_tags.column_dimensions["B"].width = 14
    ws_tags.freeze_panes = "A2"
    last_row = len(distinct_tags) + 1
    tbl = Table(displayName="Tags", ref=f"A1:B{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_tags.add_table(tbl)

    # ── Sheet: Section Distribution (live COUNTIFS against raw sheets) ─
    ws_sec = wb.create_sheet("Section Distribution")
    sections = sorted({(r["isic_section"], r["section_name"]) for r in isic_lookup}
                       | {("UNCLASSIFIED", "Unclassified")})
    # order by project count descending, computed once in Python purely for
    # row ORDER (the counts shown are still live formulas, not this number)
    section_project_counts = {}
    for p in projects:
        key = (p["isic_section"], p["section_name"])
        section_project_counts[key] = section_project_counts.get(key, 0) + 1
    sections.sort(key=lambda k: -section_project_counts.get(k, 0))

    sec_headers = ["isic_section", "section_name", "project_count", "file_count"]
    _style_header_row(ws_sec, 1, 4)
    for c, h in enumerate(sec_headers, start=1):
        ws_sec.cell(row=1, column=c, value=h)
    for i, (sec, sec_name) in enumerate(sections, start=2):
        ws_sec.cell(row=i, column=1, value=sec).font = BODY_FONT
        ws_sec.cell(row=i, column=2, value=sec_name).font = BODY_FONT
        proj_cnt = ws_sec.cell(
            row=i, column=3,
            value=f"=COUNTIF('Project Classifications'!$F$2:$F${MAX_ROW},A{i})",
        )
        file_cnt = ws_sec.cell(
            row=i, column=4,
            value=f"=COUNTIF('File Classifications'!$F$2:$F${MAX_ROW},A{i})",
        )
        for cell in (ws_sec.cell(row=i, column=1), ws_sec.cell(row=i, column=2), proj_cnt, file_cnt):
            cell.border = BORDER
            if cell.font is None:
                cell.font = BODY_FONT
        proj_cnt.font = BODY_FONT
        file_cnt.font = BODY_FONT
    ws_sec.column_dimensions["A"].width = 14
    ws_sec.column_dimensions["B"].width = 55
    ws_sec.column_dimensions["C"].width = 14
    ws_sec.column_dimensions["D"].width = 12
    ws_sec.freeze_panes = "A2"
    tbl = Table(displayName="SectionDistribution", ref=f"A1:D{len(sections)+1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_sec.add_table(tbl)

    # ── Sheet: Division Distribution (live COUNTIFS/AVERAGEIFS) ────────
    ws_div = wb.create_sheet("Division Distribution")
    divisions = sorted({(r["isic_section"], r["isic_division"], r["division_name"])
                         for r in isic_lookup} | {("UNCLASSIFIED", "00", "Unclassified")})
    division_project_counts = {}
    for p in projects:
        key = (p["isic_section"], p["isic_division"], p["division_name"])
        division_project_counts[key] = division_project_counts.get(key, 0) + 1
    divisions.sort(key=lambda k: -division_project_counts.get(k, 0))

    div_headers = ["isic_section", "isic_division", "division_name",
                   "project_count", "file_count", "avg_confidence"]
    _style_header_row(ws_div, 1, 6)
    for c, h in enumerate(div_headers, start=1):
        ws_div.cell(row=1, column=c, value=h)
    for i, (sec, div, div_name) in enumerate(divisions, start=2):
        ws_div.cell(row=i, column=1, value=sec).font = BODY_FONT
        ws_div.cell(row=i, column=2, value=div).font = BODY_FONT
        ws_div.cell(row=i, column=3, value=div_name).font = BODY_FONT
        proj_cnt = ws_div.cell(
            row=i, column=4,
            value=(f"=COUNTIFS('Project Classifications'!$F$2:$F${MAX_ROW},A{i},"
                   f"'Project Classifications'!$H$2:$H${MAX_ROW},B{i})"),
        )
        file_cnt = ws_div.cell(
            row=i, column=5,
            value=(f"=COUNTIFS('File Classifications'!$F$2:$F${MAX_ROW},A{i},"
                   f"'File Classifications'!$G$2:$G${MAX_ROW},B{i})"),
        )
        avg_conf = ws_div.cell(
            row=i, column=6,
            value=(f"=IFERROR(AVERAGEIFS('Project Classifications'!$J$2:$J${MAX_ROW},"
                   f"'Project Classifications'!$F$2:$F${MAX_ROW},A{i},"
                   f"'Project Classifications'!$H$2:$H${MAX_ROW},B{i}),0)"),
        )
        avg_conf.number_format = "0.00"
        for cell in (ws_div.cell(row=i, column=1), ws_div.cell(row=i, column=2),
                     ws_div.cell(row=i, column=3), proj_cnt, file_cnt, avg_conf):
            cell.border = BORDER
            cell.font = BODY_FONT
    ws_div.column_dimensions["A"].width = 14
    ws_div.column_dimensions["B"].width = 14
    ws_div.column_dimensions["C"].width = 55
    ws_div.column_dimensions["D"].width = 14
    ws_div.column_dimensions["E"].width = 11
    ws_div.column_dimensions["F"].width = 14
    ws_div.freeze_panes = "A2"
    tbl = Table(displayName="DivisionDistribution", ref=f"A1:F{len(divisions)+1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_div.add_table(tbl)

    # ── Sheet: Overview (title + live summary formulas) ────────────────
    ws_ov = wb.create_sheet("Overview")
    wb.move_sheet("Overview", offset=-(len(wb.sheetnames) - 1))  # make it the first sheet

    ws_ov.cell(row=1, column=1, value="Seeding QDArchive — Part 2 Classification Results").font = TITLE_FONT
    ws_ov.cell(row=2, column=1,
               value=f"{STUDENT_NAME} ({STUDENT_ID}) — Applied Software Engineering Project (10 ECTS)"
               ).font = SUBTITLE_FONT
    ws_ov.cell(row=3, column=1,
               value=f"Generated {datetime.now(timezone.utc).strftime('%d %B %Y')} from "
                     f"{Path(db_path).name}").font = SUBTITLE_FONT

    metrics = [
        ("Total projects",
         "=COUNTA('Project Classifications'!$A$2:$A$" + str(MAX_ROW) + ")"),
        ("Projects classified (ISIC section assigned)",
         "=COUNTIFS('Project Classifications'!$F$2:$F$" + str(MAX_ROW) + ',"<>UNCLASSIFIED",'
         "'Project Classifications'!$F$2:$F$" + str(MAX_ROW) + ',"<>")'),
        ("Projects unclassified",
         "=COUNTIF('Project Classifications'!$F$2:$F$" + str(MAX_ROW) + ',"UNCLASSIFIED")'),
        ("Total files classified", "=COUNTA('File Classifications'!$A$2:$A$" + str(MAX_ROW) + ")"),
        ("Files classified from own content",
         "=COUNTIF('File Classifications'!$J$2:$J$" + str(MAX_ROW) + ',"RULE_BASED_KEYWORDS")'),
        ("Files inherited from parent project (no extractable text)",
         "=COUNTIF('File Classifications'!$J$2:$J$" + str(MAX_ROW) + ',"NO_TEXT_CONTENT")'),
        ("Distinct searchable tags", "=COUNTA('Tags'!$A$2:$A$" + str(MAX_ROW) + ")"),
        ("ISIC sections represented",
         "=COUNTA('Section Distribution'!$A$2:$A$" + str(MAX_ROW) + ")"),
        ("ISIC divisions represented",
         "=COUNTA('Division Distribution'!$A$2:$A$" + str(MAX_ROW) + ")"),
    ]
    start = 5
    ws_ov.cell(row=start, column=1, value="Metric").font = HEADER_FONT
    ws_ov.cell(row=start, column=1).fill = HEADER_FILL
    ws_ov.cell(row=start, column=2, value="Value").font = HEADER_FONT
    ws_ov.cell(row=start, column=2).fill = HEADER_FILL
    for r, (label, formula) in enumerate(metrics, start=start + 1):
        lc = ws_ov.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.border = BORDER
        vc = ws_ov.cell(row=r, column=2, value=formula)
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = Alignment(horizontal="center")

    note_row = start + len(metrics) + 2
    ws_ov.cell(row=note_row, column=1,
               value="Methodology: rule-based keyword classifier scored against the complete "
                     "official ISIC Rev. 5 taxonomy (22 sections, 87 divisions), using both "
                     "project metadata and extracted file content. See classification/classifier.py "
                     "and 23123639_SQ26_Part2_Report.pdf for full methodology and technical "
                     "challenges.").font = Font(name=FONT_NAME, size=9, italic=True, color="777777")
    ws_ov.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws_ov.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    ws_ov.row_dimensions[note_row].height = 45

    ws_ov.column_dimensions["A"].width = 55
    ws_ov.column_dimensions["B"].width = 20

    wb.save(out_path)
    print(f"[XLSX] Workbook written to {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("db_path", help="Path to the classified database")
    parser.add_argument("--out", default="23123639_SQ26_Part2_Classification.xlsx")
    args = parser.parse_args()
    build_workbook(args.db_path, args.out)
